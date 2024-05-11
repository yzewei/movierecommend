
import openpyxl.writer.excel
from django.shortcuts import render, redirect ,HttpResponse
from django.utils.http import urlquote
from django.views.decorators.csrf import csrf_protect


from adminapp import models
from django.utils.safestring import mark_safe

# Create your views here.
# 后台数据管理=登录页面
from adminapp.pagination import Pagination


def login(request):
    if request.method == 'GET':
        return render(request, 'admin_login.html')
    else:
        # print(requset.POST)
        manager_id = request.POST.get("manager_id")
        password = request.POST.get("password")
        user_list = models.ManagerInfo.objects.filter(manager_id=manager_id).values("manager_id", "password").first()
        try:
            manager_id = user_list["manager_id"]
            passw = user_list["password"]
        except:
            error_name = '没有此用户'
            return render(request, 'admin_login.html', {'error_name': error_name})
        if manager_id == manager_id and password == passw:
            # return HttpResponse("登录成功")
            request.session["admin"] = manager_id
            return redirect('/manager/index')
        elif password != passw:
            error_password = '密码错误'
            return render(request, 'admin_login.html', {"error_password": error_password})
        return render(request, 'admin_login.html')


def index(request):
    admin = request.session.get("admin")
    if not admin:
        return redirect('/manager/login')
    data_dict = {}
    search_data = request.GET.get("q", "")
    if search_data:
        data_dict["uid__contains"] = search_data
    queryset = models.UserInfo.objects.filter(**data_dict)
    # 2.实例化分页对象
    page_object = Pagination(request, queryset)

    context = {
        "queryset": page_object.page_queryset,  # 分完页的数据
        "page_string": page_object.html()  # 生成页码
    }
    return render(request, "admin_index.html", context)


# 默认首页显示用户管理
# def index(request):
#     data_dict = {}
#     search_data = request.GET.get("q", "")
#     if search_data:
#         data_dict["uid__contains"] = search_data
#     # 分页
#     page = int(request.GET.get("page", 1))
#     page_size = 10
#     start = (page - 1) * page_size
#     end = page * page_size
#
#     # 获取用户表中的分页数据
#     user_list = models.UserInfo.objects.filter(**data_dict)[start:end]
#     user_count = models.UserInfo.objects.filter(**data_dict).count()
#
#     # 总页码数量
#     user_count_page, div = divmod(user_count, page_size)
#
#     # 如果余数不等于0，页数加一
#     if div:
#         user_count_page += 1
#
#     # 计算前后五页，并显示
#     plus = 5
#     if user_count_page <= 2 * plus + 1:
#         # 数据库页数少于110条
#         start_page = 1
#         end_page = user_count_page
#     else:
#         # 数据库页数较多
#         # 当前页小于5时
#         if page <= plus:
#             start_page = 1
#             end_page = 2 * plus + 1
#         else:
#             start_page = page - plus
#             end_page = page + plus + 1
#     # 存放页码
#     page_list1 = []
#     for i in range(start_page, end_page):
#         if i == page:
#             ele = ' <li class="active"><a href="/manager/index?page={}">{}</a></li>'.format(i, i)
#         else:
#             ele = ' <li><a href="/manager/index?page={}">{}</a></li>'.format(i, i)
#         page_list1.append(ele)
#     page_list = mark_safe("".join(page_list1))
#     print(page_list)
#     return render(request, "admin_index.html",
#                   {"user_list": user_list,
#                    "page_list": page_list,
#                    "search_data": search_data
#                    })


# 添加用户
def user_add(request):
    if request.method == 'GET':
        return render(request, "user_add.html")
    # 获取post传递的数据，写入数据库中
    uid = request.POST.get("username")
    password = request.POST.get("password")
    tag = request.POST.get("tag")
    like = request.POST.get("like")
    models.UserInfo.objects.create(uid=uid, password=password, tag=tag, likes=like)
    # 回到管理首页
    return redirect("/manager/index")


# 删除用户
def user_del(request):
    # http://127.0.0.1:8000/manager/user_del/?username=uid
    username = request.GET.get("uid")
    models.UserInfo.objects.filter(uid=username).delete()
    return redirect("/manager/index")


# 编辑用户
def user_edit(request, uid):
    if request.method == "GET":
        row_object = models.UserInfo.objects.filter(uid=uid).first()
        return render(request, "user_edit.html", {"row_object": row_object})
    username = request.POST.get("username")
    password = request.POST.get("password")
    tag = request.POST.get("tag")
    like = request.POST.get("like")
    models.UserInfo.objects.filter(uid=uid).update(uid=username, password=password,
                                                   tag=tag, likes=like)
    return redirect("/manager/index")


# 电影数据-管理页面
def index_movie(request):
    data_dict = {}
    search_data = request.GET.get("q", "")
    if search_data:
        data_dict["name__contains"] = search_data
    queryset = models.MoviesDetail.objects.filter(**data_dict)
    # 2.实例化分页对象
    page_object = Pagination(request, queryset)

    context = {
        "queryset": page_object.page_queryset,  # 分完页的数据
        "page_string": page_object.html()  # 生成页码
    }
    return render(request, "admin_index_movie.html", context)


# 电影数据-管理页面
# def index_movie(request):
#     data_dict = {}
#     search_data = request.GET.get("q", "")
#
#     if search_data:
#         data_dict["name__contains"] = search_data
#     # 分页
#     page = int(request.GET.get("page", 1))
#     page_size = 10
#     start = (page - 1) * page_size
#     end = page * page_size
#
#     # 获取用户表中的分页数据
#     movie_list = models.MoviesDetail.objects.filter(**data_dict)[start:end]
#     movie_count = models.MoviesDetail.objects.filter(**data_dict).count()
#
#     # 总页码数量
#     movie_count_page, div = divmod(movie_count, page_size)
#
#     # 如果余数不等于0，页数加一
#     if div:
#         movie_count_page += 1
#
#     # 计算前后五页，并显示
#     plus = 5
#     if movie_count_page <= 2 * plus + 1:
#         # 数据库页数少于110条
#         start_page = 1
#         end_page = movie_count_page
#     else:
#         # 数据库页数较多
#         # 当前页小于5时
#         if page <= plus:
#             start_page = 1
#             end_page = 2 * plus + 1
#         else:
#             start_page = page - plus
#             end_page = page + plus + 1
#
#     # print(start_page, end_page)
#     # 存放页码
#     page_list1 = []
#     for i in range(start_page, end_page):
#         if i == page:
#             ele = ' <li class="active"><a href="/manager/index_movie?page={}">{}</a></li>'.format(i, i)
#         else:
#             ele = ' <li><a href="/manager/index_movie?page={}">{}</a></li>'.format(i, i)
#         page_list1.append(ele)
#     page_list = mark_safe("".join(page_list1))
#
#     return render(request, "admin_index_movie.html",
#                   {"movie_list": movie_list,
#                    "page_list": page_list,
#                    "search_data": search_data
#                    })


# 电影数据添加
def movie_add(request):
    if request.method == 'GET':
        return render(request, "movie_add.html")
    # 获取post传递的数据，写入数据库中
    moviename1 = request.POST.get("moviename")
    directors = request.POST.get("directors")
    writer = request.POST.get("writer")
    actors = request.POST.get("actors")
    rate = request.POST.get("rate")
    dataID = request.POST.get("dataID")

    date = request.POST.get("date")
    style1 = request.POST.get("style1")
    style2 = request.POST.get("style2")
    style3 = request.POST.get("style3")
    introduction = request.POST.get("introduction")
    pic = request.POST.get("pic")
    english_name = request.POST.get("english_name")

    url = request.POST.get("url")
    language = request.POST.get("language")
    country = request.POST.get("country")
    duration = request.POST.get("duration")

    models.MoviesDetail.objects.create(name=moviename1, directors=directors,
                                       writer=writer, actors=actors, rate=rate,
                                       date=date,style1=style1,style2=style2,style3=style3,
                                       pic=pic,url=url,language=language,country=country,duration=duration,
                                       english_name=english_name,dataid=dataID,introduction=introduction)
    # 回到电影管理首页
    return redirect("/manager/index_movie")


# 电影数据删除
def movie_del(request):
    # http://127.0.0.1:8000/manager/movie_del/?dataid=dataid
    dataid = request.GET.get("dataid")
    print(dataid)
    models.MoviesDetail.objects.filter(dataid=dataid).delete()
    return redirect("/manager/index_movie")


# 电影数据编辑
def movie_edit(request, dataid):
    if request.method == "GET":
        row_object = models.MoviesDetail.objects.filter(dataid=dataid).first()
        return render(request, "movie_edit.html", {"row_object": row_object})
    moviename1 = request.POST.get("moviename")
    directors = request.POST.get("directors")
    writer = request.POST.get("writer")
    actors = request.POST.get("actors")
    rate = request.POST.get("rate")
    date = request.POST.get("date")
    dataid1 = request.POST.get("dataid")
    models.MoviesDetail.objects.filter(dataid=dataid).update(
        name=moviename1, directors=directors, writer=writer, actors=actors, rate=rate, dataid=dataid1, date=date)
    return redirect("/manager/index_movie")


def index_admin(request):
    data_dict = {}
    search_data = request.GET.get("q", "")
    if search_data:
        data_dict["manager_id__contains"] = search_data
    # 筛选数据
    queryset = models.ManagerInfo.objects.filter(**data_dict)
    # 实例化分页对象
    page_object = Pagination(request, queryset)

    context = {
        "queryset": page_object.page_queryset,  # 分完页的数据
        "page_string": page_object.html()  # 生成页码
    }
    return render(request, "admin_index_admin.html", context)


# 管理员数据-管理页面
# def index_admin(request):
#     data_dict = {}
#     search_data = request.GET.get("q", "")
#
#     if search_data:
#         data_dict["manager_id__contains"] = search_data
#     # 分页
#     page = int(request.GET.get("page", 1))
#     page_size = 10
#     start = (page - 1) * page_size
#     end = page * page_size
#
#     # 获取用户表中的分页数据
#     manager_list = models.ManagerInfo.objects.filter(**data_dict)[start:end]
#     manager_count = models.ManagerInfo.objects.filter(**data_dict).count()
#
#     # 总页码数量
#     manager_count_page, div = divmod(manager_count, page_size)
#
#     # 如果余数不等于0，页数加一
#     if div:
#         manager_count_page += 1
#
#     # 计算前后五页，并显示
#     plus = 5
#     if manager_count_page <= 2 * plus + 1:
#         # 数据库页数少于110条
#         start_page = 1
#         end_page = manager_count_page
#     else:
#         # 数据库页数较多
#         # 当前页小于5时
#         if page <= plus:
#             start_page = 1
#             end_page = 2 * plus + 1
#         else:
#             start_page = page - plus
#             end_page = page + plus + 1
#     # 存放页码
#     page_list1 = []
#     for i in range(start_page, end_page):
#         if i == page:
#             ele = ' <li class="active"><a href="/manager/index_admin?page={}">{}</a></li>'.format(i, i)
#         else:
#             ele = ' <li><a href="/manager/index_admin?page={}">{}</a></li>'.format(i, i)
#         page_list1.append(ele)
#     page_list = mark_safe("".join(page_list1))
#
#     return render(request, "admin_index_admin.html",
#                   {"manager_list": manager_list,
#                    "page_list": page_list,
#                    "search_data": search_data})


# 管理员数据添加
def admin_add(request):
    if request.method == 'GET':
        return render(request, "admin_add.html")
    # 获取post传递的数据，写入数据库中
    manager_id = request.POST.get("manager_id")
    password = request.POST.get("password")

    models.ManagerInfo.objects.create(manager_id=manager_id, password=password)
    # 回到电影管理首页
    return redirect("/manager/index_admin")


# 管理员数据删除
def admin_del(request):
    # http://127.0.0.1:8000/manager/admin_del/?manager_id=manager_id
    manager_id = request.GET.get("manager_id")
    models.ManagerInfo.objects.filter(manager_id=manager_id).delete()
    return redirect("/manager/index_admin")


# 管理员数据编辑
def admin_edit(request, manager_id):
    if request.method == "GET":
        row_object = models.ManagerInfo.objects.filter(manager_id=manager_id).first()
        return render(request, "admin_edit.html", {"row_object": row_object})
    manager_id1 = request.POST.get("manager_id")
    password = request.POST.get("password")
    print(manager_id1, password)
    models.ManagerInfo.objects.filter(manager_id=manager_id).update(manager_id=manager_id1, password=password)
    return redirect("/manager/index_admin")


# 用户数据可视化
def user_show(request):
    return render(request, "admin_user_show.html")


def echarts(request):
    return render(request, "result.html")


# 注销 清除cookie
def adminlogout(request):
    # request.session.clear()
    del request.session["admin"]
    return redirect('/manager/login')


from openpyxl import load_workbook

@csrf_protect
# 上传用户数据文件Excel
def user_multi(request):
    # 获取上传对象
    file_object = request.FILES.get("execl")
    # 由openpyxl读取文件内容
    wb = load_workbook(file_object)
    # 读取第一页sheet的内容
    sheet = wb.worksheets[0]
    # 读取第一行第一列
    # cell = sheet.cell(1, 1)
    # 循环读取每一行数据
    for row in sheet.iter_rows(min_row=2):
        uid = row[0].value
        password = row[1].value
        tag = row[2].value
        likes = row[3].value
        date = row[4].value
        exists =models.UserInfo.objects.filter(uid=uid).exists()
        # 判断用户是否存在
        if not exists:
            try:
                models.UserInfo.objects.create(uid=uid ,password=password ,tag=tag ,likes=likes ,createdate=date)
            except:
                return HttpResponse("文件内容有误，请检查后再上传")
    return redirect('/manager/index')



import os
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, colors, PatternFill
from openpyxl.utils import get_column_letter
from io import BytesIO
import datetime
FILE_PATH = 'C:/Users/admin/Downloads/'
# 下载用户数据文件Excel
def user_export(request):
    wb = Workbook()  # 生成一个工作簿（即一个Excel文件）
    wb.encoding = 'utf-8'
    sheet1 = wb.active  # 获取第一个工作表（sheet1）
    sheet1.title = 'user_info'  # 给工作表1设置标题
    row_one = ['uid', 'password', 'tag' ,'likes', 'createDate', 'updateDate']
    for i in range(1, len(row_one) + 1):  # 从第一行开始写，因为Excel文件的行号是从1开始，列号也是从1开始
        # 从row=1，column=1开始写，即将row_one的数据依次写入第一行
        sheet1.cell(row=1, column=i).value = row_one[i - 1]
    all_obj = models.UserInfo.objects.all()
    for obj in all_obj:
        max_row = sheet1.max_row + 1  # 获取到工作表的最大行数并加1
        obj_info = [obj.uid ,obj.password ,obj.tag ,obj.likes ,obj.createdate ,obj.updatedate ]
        for x in range(1, len(obj_info) + 1):  # 将每一个对象的所有字段的信息写入一行内
            sheet1.cell(row=max_row, column=x).value = obj_info[x - 1]
    # 准备写入到IO中
    output = BytesIO()
    wb.save(output)  # 将Excel文件内容保存到IO中
    output.seek(0)  # 重新定位到开始
    # 设置HttpResponse的类型
    response = HttpResponse(output.getvalue(), content_type='application/vnd.ms-excel')
    ctime = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    file_name = 'user_AllInfo%s.xls' % ctime  # 给文件名中添加日期时间
    file_name = urlquote(file_name)  # 使用urlquote()方法解决中文无法使用的问题
    response['Content-Disposition'] = 'attachment; filename=%s' % file_name
    return response


@csrf_protect
# 上传电影数据Excel
def movie_multi(request):
    # 获取上传对象
    file_object = request.FILES.get("execl")
    # 由openpyxl读取文件内容
    wb = load_workbook(file_object)
    # 读取第一页sheet的内容
    sheet = wb.worksheets[0]
    # 读取第一行第一列
    # cell = sheet.cell(1, 1)
    # 循环读取每一行数据
    for row in sheet.iter_rows(min_row=2):
        name = row[0].value
        english_name = row[1].value
        directors = row[2].value
        writer = row[3].value
        actors = row[4].value
        rate = row[5].value
        style1 = row[6].value
        style2 = row[7].value
        style3 = row[8].value
        country = row[9].value
        language = row[10].value
        date = row[11].value
        duration = row[12].value
        introduction = row[13].value
        dataid = row[14].value
        url = row[15].value
        pic = row[16].value
        exists = models.MoviesDetail.objects.filter(dataid=dataid).exists()
        # 判断用户是否存在
        if not exists:
            try:
                models.MoviesDetail.objects.create(name=name ,english_name=english_name ,directors=directors
                                                   ,writer=writer,
                                                   actors=actors ,rate=rate ,style1=style1 ,style2=style2
                                                   ,style3=style3,
                                                   country=country ,language=language ,date=date ,duration=duration
                                                   ,introduction=introduction,
                                                   dataid=dataid ,url=url ,pic=pic)
            except:
                return HttpResponse("文件内容有误，请检查后再上传")
        return redirect('/manager/index_movie')


def movie_export(request):
    wb = Workbook()  # 生成一个工作簿（即一个Excel文件）
    wb.encoding = 'utf-8'
    sheet1 = wb.active  # 获取第一个工作表（sheet1）
    sheet1.title = 'movie_info'  # 给工作表1设置标题
    row_one = ['name','english_name','directors','writer','actors','rate',
               'style1','style2','style3','country','language','date','duration',
               'introduction','dataID','url','pic']
    for i in range(1, len(row_one) + 1):  # 从第一行开始写，因为Excel文件的行号是从1开始，列号也是从1开始
        # 从row=1，column=1开始写，即将row_one的数据依次写入第一行
        sheet1.cell(row=1, column=i).value = row_one[i - 1]
    all_obj = models.MoviesDetail.objects.all()[0:1000]
    for obj in all_obj:
        max_row = sheet1.max_row + 1  # 获取到工作表的最大行数并加1
        obj_info = [obj.name,obj.english_name,obj.directors,obj.writer,obj.actors,obj.rate,
                    obj.style1,obj.style2,obj.style3,obj.country,obj.language,obj.date,
                    obj.duration,obj.introduction,obj.dataid,obj.url,obj.pic]
        for x in range(1, len(obj_info) + 1):  # 将每一个对象的所有字段的信息写入一行内
            sheet1.cell(row=max_row, column=x).value = obj_info[x - 1]
    # 准备写入到IO中
    output = BytesIO()
    wb.save(output)  # 将Excel文件内容保存到IO中
    output.seek(0)  # 重新定位到开始
    # 设置HttpResponse的类型
    response = HttpResponse(output.getvalue(), content_type='application/vnd.ms-excel')
    ctime = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    file_name = 'movie_AllInfo%s.xls' % ctime  # 给文件名中添加日期时间
    file_name = urlquote(file_name)  # 使用urlquote()方法解决中文无法使用的问题
    response['Content-Disposition'] = 'attachment; filename=%s' % file_name
    return response

